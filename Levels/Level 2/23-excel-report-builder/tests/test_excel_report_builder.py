"""
Comprehensive test suite for Excel Report Builder.

This module contains unit tests for all core functionality including
Excel generation, formatting, chart creation, and data processing.
"""

import os
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import pandas as pd
import pytest
from openpyxl import load_workbook

from excel_report_builder.core import ExcelReportBuilder
from excel_report_builder.utils import (
    DataProcessor,
    ChartGenerator,
    validate_output_path,
    get_column_letter,
    format_file_size
)


class TestExcelReportBuilder:
    """Test cases for ExcelReportBuilder class."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()
        self.output_file = os.path.join(self.temp_dir, "test_report.xlsx")
        self.builder = ExcelReportBuilder(self.output_file)

    def teardown_method(self):
        """Clean up test fixtures."""
        # Clean up temporary files
        if os.path.exists(self.output_file):
            os.remove(self.output_file)

    def test_initialization(self):
        """Test ExcelReportBuilder initialization."""
        assert self.builder.output_file == Path(self.output_file)
        assert self.builder.workbook is not None
        assert len(self.builder.sheets) == 0

    def test_load_csv_data_success(self):
        """Test successful CSV data loading."""
        # Create test CSV file
        test_data = pd.DataFrame({
            'Name': ['Alice', 'Bob', 'Charlie'],
            'Age': [25, 30, 35],
            'Salary': [50000, 60000, 70000]
        })
        csv_file = os.path.join(self.temp_dir, "test_data.csv")
        test_data.to_csv(csv_file, index=False)

        # Test loading
        df = self.builder.load_csv_data(csv_file, "TestSheet")

        assert len(df) == 3
        assert list(df.columns) == ['Name', 'Age', 'Salary']
        assert "TestSheet" in self.builder.sheets

    def test_load_csv_data_file_not_found(self):
        """Test CSV loading with non-existent file."""
        with pytest.raises(FileNotFoundError):
            self.builder.load_csv_data("nonexistent.csv")

    def test_load_csv_data_empty_file(self):
        """Test CSV loading with empty file."""
        # Create empty CSV file
        empty_csv = os.path.join(self.temp_dir, "empty.csv")
        with open(empty_csv, 'w') as f:
            f.write("")

        with pytest.raises(pd.errors.EmptyDataError):
            self.builder.load_csv_data(empty_csv)

    def test_create_sheet_from_dataframe(self):
        """Test creating sheet from DataFrame."""
        test_data = pd.DataFrame({
            'Product': ['Widget A', 'Widget B', 'Widget C'],
            'Sales': [100, 200, 150],
            'Profit': [20, 40, 30]
        })

        ws = self.builder.create_sheet_from_dataframe(test_data, "Products")

        assert "Products" in self.builder.sheets
        assert ws.max_row == 4  # Header + 3 data rows
        assert ws.max_column == 3

    def test_add_formulas(self):
        """Test adding formulas to worksheet."""
        # Create test data
        test_data = pd.DataFrame({
            'A': [1, 2, 3],
            'B': [4, 5, 6]
        })
        self.builder.create_sheet_from_dataframe(test_data, "TestSheet")

        # Add formulas
        formulas = {
            'A5': '=SUM(A2:A4)',
            'B5': '=AVERAGE(B2:B4)'
        }
        self.builder.add_formulas("TestSheet", formulas)

        ws = self.builder.sheets["TestSheet"]
        assert ws['A5'].value == '=SUM(A2:A4)'
        assert ws['B5'].value == '=AVERAGE(B2:B4)'

    def test_create_chart(self):
        """Test chart creation."""
        # Create test data
        test_data = pd.DataFrame({
            'Product': ['A', 'B', 'C'],
            'Sales': [100, 200, 150]
        })
        self.builder.create_sheet_from_dataframe(test_data, "TestSheet")

        # Create chart
        self.builder.create_chart(
            "TestSheet",
            "bar",
            "B2:B4",
            "A2:A4",
            "Sales Chart"
        )

        ws = self.builder.sheets["TestSheet"]
        assert len(ws._charts) == 1

    def test_create_chart_detects_headers(self):
        """chart.add_data should only treat first row as headers when provided."""
        test_data = pd.DataFrame({
            'Product': ['A', 'B', 'C'],
            'Sales': [100, 200, 150]
        })
        self.builder.create_sheet_from_dataframe(test_data, "HeaderTest")

        with patch('excel_report_builder.core.BarChart.add_data') as mock_add_data:
            self.builder.create_chart(
                "HeaderTest",
                "bar",
                "B2:B4",
                "A2:A4",
                "No Header Range"
            )
        assert mock_add_data.call_args.kwargs["titles_from_data"] is False

        with patch('excel_report_builder.core.BarChart.add_data') as mock_add_data:
            self.builder.create_chart(
                "HeaderTest",
                "bar",
                "B1:B4",
                "A2:A4",
                "Header Included"
            )
        assert mock_add_data.call_args.kwargs["titles_from_data"] is True

    def test_add_conditional_formatting(self):
        """Test conditional formatting."""
        # Create test data
        test_data = pd.DataFrame({
            'Values': [10, 20, 30, 40, 50]
        })
        self.builder.create_sheet_from_dataframe(test_data, "TestSheet")

        # Add conditional formatting
        self.builder.add_conditional_formatting(
            "TestSheet",
            "A2:A6",
            "greater_than",
            25,
            "FFFF00"
        )

        ws = self.builder.sheets["TestSheet"]
        assert len(ws.conditional_formatting.cf_rules) == 1

    def test_create_summary_sheet(self):
        """Test summary sheet creation."""
        # Create test data
        test_data = pd.DataFrame({
            'Name': ['Alice', 'Bob'],
            'Value': [100, 200]
        })
        self.builder.create_sheet_from_dataframe(test_data, "Data")

        # Create summary sheet
        self.builder.create_summary_sheet("Data", "Summary")

        assert "Summary" in self.builder.sheets
        summary_ws = self.builder.sheets["Summary"]
        assert summary_ws.max_row >= 3  # Headers + data

    def test_save_workbook(self):
        """Test workbook saving."""
        # Create test data
        test_data = pd.DataFrame({'A': [1, 2, 3]})
        self.builder.create_sheet_from_dataframe(test_data, "TestSheet")

        # Save workbook
        self.builder.save_workbook()

        assert os.path.exists(self.output_file)

        # Verify file can be loaded
        wb = load_workbook(self.output_file)
        assert "TestSheet" in wb.sheetnames

    def test_get_sheet_names(self):
        """Test getting sheet names."""
        self.builder.create_sheet_from_dataframe(
            pd.DataFrame({'A': [1]}), "Sheet1"
        )
        self.builder.create_sheet_from_dataframe(
            pd.DataFrame({'B': [2]}), "Sheet2"
        )

        sheet_names = self.builder.get_sheet_names()
        assert "Sheet1" in sheet_names
        assert "Sheet2" in sheet_names

    def test_get_sheet_info(self):
        """Test getting sheet information."""
        test_data = pd.DataFrame({
            'A': [1, 2, 3],
            'B': [4, 5, 6]
        })
        self.builder.create_sheet_from_dataframe(test_data, "TestSheet")

        info = self.builder.get_sheet_info("TestSheet")
        assert info["name"] == "TestSheet"
        assert info["max_row"] == 4  # Header + 3 data rows
        assert info["max_column"] == 2


class TestDataProcessor:
    """Test cases for DataProcessor class."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        """Clean up test fixtures."""
        # Clean up temporary files
        for file in os.listdir(self.temp_dir):
            os.remove(os.path.join(self.temp_dir, file))
        os.rmdir(self.temp_dir)

    def test_validate_csv_file_valid(self):
        """Test CSV file validation with valid file."""
        # Create valid CSV file
        csv_file = os.path.join(self.temp_dir, "valid.csv")
        pd.DataFrame({'A': [1, 2, 3]}).to_csv(csv_file, index=False)

        assert DataProcessor.validate_csv_file(csv_file) is True

    def test_validate_csv_file_nonexistent(self):
        """Test CSV file validation with non-existent file."""
        assert DataProcessor.validate_csv_file("nonexistent.csv") is False

    def test_validate_csv_file_wrong_extension(self):
        """Test CSV file validation with wrong file extension."""
        # Create non-CSV file
        txt_file = os.path.join(self.temp_dir, "test.txt")
        with open(txt_file, 'w') as f:
            f.write("not csv data")

        assert DataProcessor.validate_csv_file(txt_file) is False

    def test_clean_dataframe(self):
        """Test DataFrame cleaning."""
        # Create DataFrame with empty rows and columns
        df = pd.DataFrame({
            'A': [1, 2, None, 4],
            'B': [None, None, None, None],
            'C': [5, 6, 7, 8]
        })

        cleaned = DataProcessor.clean_dataframe(df)

        assert len(cleaned) == 3  # Removed empty row
        assert len(cleaned.columns) == 2  # Removed empty column B

    def test_detect_numeric_columns(self):
        """Test numeric column detection."""
        df = pd.DataFrame({
            'numeric_int': [1, 2, 3],
            'numeric_float': [1.1, 2.2, 3.3],
            'string_numeric': ['1', '2', '3'],
            'text': ['a', 'b', 'c']
        })

        numeric_cols = DataProcessor.detect_numeric_columns(df)

        assert 'numeric_int' in numeric_cols
        assert 'numeric_float' in numeric_cols
        assert 'string_numeric' in numeric_cols
        assert 'text' not in numeric_cols

    def test_detect_date_columns(self):
        """Test date column detection."""
        df = pd.DataFrame({
            'date_col': pd.date_range('2024-01-01', periods=3),
            'string_date': ['2024-01-01', '2024-01-02', '2024-01-03'],
            'text': ['a', 'b', 'c']
        })

        date_cols = DataProcessor.detect_date_columns(df)

        assert 'date_col' in date_cols
        assert 'string_date' in date_cols
        assert 'text' not in date_cols

    def test_calculate_summary_stats(self):
        """Test summary statistics calculation."""
        df = pd.DataFrame({
            'A': [1, 2, 3, 4, 5],
            'B': ['a', 'b', 'c', 'd', 'e'],
            'C': [1.1, 2.2, 3.3, 4.4, 5.5]
        })

        stats = DataProcessor.calculate_summary_stats(df)

        assert stats['total_rows'] == 5
        assert stats['total_columns'] == 3
        assert 'A' in stats['numeric_columns']
        assert 'C' in stats['numeric_columns']
        assert 'B' not in stats['numeric_columns']

    def test_prepare_chart_data(self):
        """Test chart data preparation."""
        df = pd.DataFrame({
            'x': ['A', 'B', 'C'],
            'y': [10, 20, 30]
        })

        x_values, y_values = DataProcessor.prepare_chart_data(df, 'x', 'y')

        assert x_values == ['A', 'B', 'C']
        assert y_values == [10.0, 20.0, 30.0]

    def test_prepare_chart_data_invalid_columns(self):
        """Test chart data preparation with invalid columns."""
        df = pd.DataFrame({'A': [1, 2, 3]})

        with pytest.raises(ValueError):
            DataProcessor.prepare_chart_data(df, 'nonexistent', 'A')


class TestChartGenerator:
    """Test cases for ChartGenerator class."""

    def test_get_chart_config_bar(self):
        """Test bar chart configuration."""
        config = ChartGenerator.get_chart_config('bar', 'Test Chart', 'X', 'Y')

        assert config['type'] == 'bar'
        assert config['title'] == 'Test Chart'
        assert config['x_label'] == 'X'
        assert config['y_label'] == 'Y'

    def test_get_chart_config_line(self):
        """Test line chart configuration."""
        config = ChartGenerator.get_chart_config('line', 'Test Chart')

        assert config['type'] == 'line'
        assert config['markers'] is True
        assert config['smooth'] is True

    def test_get_chart_config_pie(self):
        """Test pie chart configuration."""
        config = ChartGenerator.get_chart_config('pie', 'Test Chart')

        assert config['type'] == 'pie'
        assert config['show_legend'] is True
        assert config['show_percent'] is True

    def test_get_chart_config_invalid_type(self):
        """Test chart configuration with invalid type."""
        with pytest.raises(ValueError):
            ChartGenerator.get_chart_config('invalid', 'Test Chart')

    def test_suggest_chart_type(self):
        """Test chart type suggestion."""
        df = pd.DataFrame({
            'category': ['A', 'B', 'C', 'D', 'E'],
            'value': [10, 20, 30, 40, 50]
        })

        chart_type = ChartGenerator.suggest_chart_type(df, 'category', 'value')
        assert chart_type in ['bar', 'line', 'pie']

    def test_get_color_scheme(self):
        """Test color scheme retrieval."""
        default_scheme = ChartGenerator.get_color_scheme('default')
        assert len(default_scheme) == 5
        assert all(color.startswith('#') for color in default_scheme)

        custom_scheme = ChartGenerator.get_color_scheme('nonexistent')
        assert custom_scheme == default_scheme  # Should fallback to default


class TestUtilityFunctions:
    """Test cases for utility functions."""

    def test_format_file_size(self):
        """Test file size formatting."""
        assert format_file_size(1024) == "1.0 KB"
        assert format_file_size(1048576) == "1.0 MB"
        assert format_file_size(1024 * 1024 * 1024) == "1.0 GB"

    def test_validate_output_path(self):
        """Test output path validation."""
        with tempfile.TemporaryDirectory() as temp_dir:
            valid_path = os.path.join(temp_dir, "test.xlsx")
            assert validate_output_path(valid_path) is True

            invalid_path = os.path.join(temp_dir, "test.txt")
            assert validate_output_path(invalid_path) is True  # Should work but warn

    def test_get_column_letter(self):
        """Test column letter conversion."""
        assert get_column_letter(1) == 'A'
        assert get_column_letter(2) == 'B'
        assert get_column_letter(26) == 'Z'
        assert get_column_letter(27) == 'AA'
        assert get_column_letter(28) == 'AB'


class TestIntegration:
    """Integration tests for the complete workflow."""

    def setup_method(self):
        """Set up test fixtures."""
        self.temp_dir = tempfile.mkdtemp()

    def teardown_method(self):
        """Clean up test fixtures."""
        # Clean up temporary files
        for file in os.listdir(self.temp_dir):
            os.remove(os.path.join(self.temp_dir, file))
        os.rmdir(self.temp_dir)

    def test_complete_workflow(self):
        """Test complete Excel report generation workflow."""
        # Create test data
        test_data = pd.DataFrame({
            'Product': ['Widget A', 'Widget B', 'Widget C'],
            'Sales': [100, 200, 150],
            'Profit': [20, 40, 30]
        })
        csv_file = os.path.join(self.temp_dir, "test_data.csv")
        test_data.to_csv(csv_file, index=False)

        output_file = os.path.join(self.temp_dir, "test_report.xlsx")

        # Create builder and generate report
        builder = ExcelReportBuilder(output_file)
        df = builder.load_csv_data(csv_file, "Sales Data")
        builder.create_sheet_from_dataframe(df, "Sales Data")

        # Add formulas
        formulas = {
            'B5': '=SUM(B2:B4)',
            'C5': '=SUM(C2:C4)'
        }
        builder.add_formulas("Sales Data", formulas)

        # Add chart
        builder.create_chart(
            "Sales Data",
            "bar",
            "B2:B4",
            "A2:A4",
            "Sales Chart"
        )

        # Add conditional formatting
        builder.add_conditional_formatting(
            "Sales Data",
            "B2:B4",
            "greater_than",
            150,
            "90EE90"
        )

        # Create summary sheet
        builder.create_summary_sheet("Sales Data", "Summary")

        # Save workbook
        builder.save_workbook()

        # Verify output
        assert os.path.exists(output_file)

        # Load and verify workbook
        wb = load_workbook(output_file)
        assert "Sales Data" in wb.sheetnames
        assert "Summary" in wb.sheetnames

        # Verify data
        ws = wb["Sales Data"]
        assert ws['A1'].value == 'Product'
        assert ws['B1'].value == 'Sales'
        assert ws['C1'].value == 'Profit'

        # Verify formulas
        assert ws['B5'].value == '=SUM(B2:B4)'
        assert ws['C5'].value == '=SUM(C2:C4)'

        # Verify chart
        assert len(ws._charts) == 1

        # Verify conditional formatting
        assert len(ws.conditional_formatting.cf_rules) == 1


if __name__ == '__main__':
    pytest.main([__file__])
