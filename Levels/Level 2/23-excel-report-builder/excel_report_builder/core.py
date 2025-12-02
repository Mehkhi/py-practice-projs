"""
Core functionality for Excel Report Builder.

This module contains the main ExcelReportBuilder class that handles
Excel file generation, formatting, and chart creation.
"""

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
try:
    from openpyxl.utils import range_boundaries
except ImportError:
    _CELL_REF_RE = re.compile(r"\$?([A-Za-z]+)\$?(\d+)")

    def _col_to_index(label: str) -> int:
        index = 0
        for char in label.upper():
            if not ("A" <= char <= "Z"):
                raise ValueError(f"Invalid column label: {label}")
            index = index * 26 + (ord(char) - 64)
        return index

    def range_boundaries(range_string: str) -> Tuple[int, int, int, int]:
        """Fallback implementation for openpyxl.utils.range_boundaries."""
        if "!" in range_string:
            _, range_part = range_string.split("!", 1)
        else:
            range_part = range_string

        parts = [part.strip() for part in range_part.split(":") if part.strip()]
        if not parts:
            raise ValueError(f"Invalid range string: {range_string}")
        if len(parts) == 1:
            start = end = parts[0]
        elif len(parts) == 2:
            start, end = parts
        else:
            raise ValueError(f"Invalid range string: {range_string}")

        match_start = _CELL_REF_RE.fullmatch(start)
        match_end = _CELL_REF_RE.fullmatch(end)
        if not (match_start and match_end):
            raise ValueError(f"Invalid range string: {range_string}")

        min_col = _col_to_index(match_start.group(1))
        min_row = int(match_start.group(2))
        max_col = _col_to_index(match_end.group(1))
        max_row = int(match_end.group(2))
        return min_col, min_row, max_col, max_row
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class ExcelReportBuilder:
    """
    Main class for building Excel reports with formatting and charts.

    This class provides methods to:
    - Load data from CSV files
    - Generate formatted Excel workbooks
    - Add charts and formulas
    - Apply conditional formatting
    - Create multiple sheets
    """

    def __init__(self, output_file: str):
        """
        Initialize the Excel Report Builder.

        Args:
            output_file: Path to the output Excel file
        """
        self.output_file = Path(output_file)
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        self.sheets: Dict[str, Worksheet] = {}

        # Default formatting styles
        self.header_style = {
            'font': Font(bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

        self.data_style = {
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def load_csv_data(self, csv_file: str, sheet_name: str = "Data") -> pd.DataFrame:
        """
        Load data from a CSV file.

        Args:
            csv_file: Path to the CSV file
            sheet_name: Name for the sheet (default: "Data")

        Returns:
            pandas DataFrame with the loaded data

        Raises:
            FileNotFoundError: If CSV file doesn't exist
            pd.errors.EmptyDataError: If CSV file is empty
        """
        try:
            logger.info(f"Loading data from {csv_file}")
            df = pd.read_csv(csv_file)

            if df.empty:
                raise pd.errors.EmptyDataError("CSV file is empty")

            # Create sheet for the data
            self.sheets[sheet_name] = self.workbook.create_sheet(title=sheet_name)

            logger.info(f"Successfully loaded {len(df)} rows from {csv_file}")
            return df

        except FileNotFoundError:
            logger.error(f"CSV file not found: {csv_file}")
            raise
        except Exception as e:
            logger.error(f"Error loading CSV file {csv_file}: {str(e)}")
            raise

    def create_sheet_from_dataframe(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        start_row: int = 1,
        start_col: int = 1
    ) -> Worksheet:
        """
        Create a worksheet from a pandas DataFrame.

        Args:
            df: pandas DataFrame
            sheet_name: Name for the worksheet
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)

        Returns:
            openpyxl Worksheet object
        """
        if sheet_name not in self.sheets:
            self.sheets[sheet_name] = self.workbook.create_sheet(title=sheet_name)

        ws = self.sheets[sheet_name]

        # Add data to worksheet
        if getattr(pd, "_IS_STUB", False):
            for row in df.to_rows(header=True, index=False):
                ws.append(row)
        else:
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

        # Apply formatting
        self._apply_data_formatting(ws, len(df), len(df.columns), start_row, start_col)

        return ws

    def _apply_data_formatting(
        self,
        ws: Worksheet,
        num_rows: int,
        num_cols: int,
        start_row: int = 1,
        start_col: int = 1
    ) -> None:
        """
        Apply formatting to the worksheet.

        Args:
            ws: Worksheet to format
            num_rows: Number of data rows
            num_cols: Number of columns
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)
        """
        # Format header row
        for col in range(start_col, start_col + num_cols):
            cell = ws.cell(row=start_row, column=col)
            for attr, value in self.header_style.items():
                setattr(cell, attr, value)

        # Format data rows
        for row in range(start_row + 1, start_row + num_rows + 1):
            for col in range(start_col, start_col + num_cols):
                cell = ws.cell(row=row, column=col)
                for attr, value in self.data_style.items():
                    setattr(cell, attr, value)

        # Auto-adjust column widths
        for col in range(start_col, start_col + num_cols):
            column_letter = ws.cell(row=1, column=col).column_letter
            max_length = 0

            for row in range(start_row, start_row + num_rows + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    def add_formulas(
        self,
        sheet_name: str,
        formulas: Dict[str, str]
    ) -> None:
        """
        Add formulas to a worksheet.

        Args:
            sheet_name: Name of the worksheet
            formulas: Dictionary mapping cell references to formulas
        """
        if sheet_name not in self.sheets:
            logger.warning(f"Sheet '{sheet_name}' not found")
            return

        ws = self.sheets[sheet_name]

        for cell_ref, formula in formulas.items():
            ws[cell_ref] = formula
            logger.info(f"Added formula to {cell_ref}: {formula}")

    def create_chart(
        self,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        categories_range: str,
        title: str,
        position: str = "E2"
    ) -> None:
        """
        Create a chart in the specified worksheet.

        Args:
            sheet_name: Name of the worksheet
            chart_type: Type of chart ('bar', 'line', 'pie')
            data_range: Range of data values (e.g., 'B2:B10')
            categories_range: Range of category labels (e.g., 'A2:A10')
            title: Chart title
            position: Position to place the chart (e.g., 'E2')
        """
        if sheet_name not in self.sheets:
            logger.warning(f"Sheet '{sheet_name}' not found")
            return

        ws = self.sheets[sheet_name]

        # Create chart based on type
        if chart_type.lower() == 'bar':
            chart = BarChart()
        elif chart_type.lower() == 'line':
            chart = LineChart()
        elif chart_type.lower() == 'pie':
            chart = PieChart()
        else:
            logger.error(f"Unsupported chart type: {chart_type}")
            return

        # Set chart properties
        chart.title = title
        chart.style = 13

        # Add data
        data = Reference(ws, range_string=data_range)
        categories = Reference(ws, range_string=categories_range)
        try:
            _, min_row, _, _ = range_boundaries(data_range)
        except ValueError:
            logger.warning(f"Invalid data range '{data_range}' for chart; treating as data without headers")
            min_row = 2  # Assume no headers if parsing fails
        titles_from_data = min_row == 1
        chart.add_data(data, titles_from_data=titles_from_data)
        chart.set_categories(categories)

        # Position chart
        ws.add_chart(chart, position)
        logger.info(f"Created {chart_type} chart '{title}' in {sheet_name}")

    def add_conditional_formatting(
        self,
        sheet_name: str,
        cell_range: str,
        rule_type: str,
        value: Union[str, int, float],
        fill_color: str = "FFFF00"
    ) -> None:
        """
        Add conditional formatting to a worksheet.

        Args:
            sheet_name: Name of the worksheet
            cell_range: Range of cells to format (e.g., 'B2:B10')
            rule_type: Type of rule ('greater_than', 'less_than', 'equal_to')
            value: Value to compare against
            fill_color: Color to apply when condition is met
        """
        if sheet_name not in self.sheets:
            logger.warning(f"Sheet '{sheet_name}' not found")
            return

        ws = self.sheets[sheet_name]

        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill

        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        if rule_type == 'greater_than':
            rule = CellIsRule(operator='greaterThan', formula=[str(value)], fill=fill)
        elif rule_type == 'less_than':
            rule = CellIsRule(operator='lessThan', formula=[str(value)], fill=fill)
        elif rule_type == 'equal_to':
            rule = CellIsRule(operator='equal', formula=[str(value)], fill=fill)
        else:
            logger.error(f"Unsupported rule type: {rule_type}")
            return

        ws.conditional_formatting.add(cell_range, rule)
        logger.info(f"Added conditional formatting to {cell_range} in {sheet_name}")

    def create_summary_sheet(
        self,
        data_sheet: str,
        summary_name: str = "Summary"
    ) -> None:
        """
        Create a summary sheet with key statistics.

        Args:
            data_sheet: Name of the data sheet
            summary_name: Name for the summary sheet
        """
        if data_sheet not in self.sheets:
            logger.warning(f"Data sheet '{data_sheet}' not found")
            return

        # Create summary sheet
        self.sheets[summary_name] = self.workbook.create_sheet(title=summary_name)
        ws = self.sheets[summary_name]

        # Add summary headers
        headers = ["Metric", "Value"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=header)
            for attr, value in self.header_style.items():
                setattr(cell, attr, value)

        # Add summary data
        summary_data = [
            ("Total Rows", f"=COUNTA({data_sheet}!A:A)-1"),
            ("Data Range", f"{data_sheet}!A1:Z1000"),
            ("Generated", "=NOW()")
        ]

        for i, (metric, value) in enumerate(summary_data, 2):
            ws.cell(row=i, column=1, value=metric)
            ws.cell(row=i, column=2, value=value)

        # Format summary sheet
        self._apply_data_formatting(ws, len(summary_data), 2)

        logger.info(f"Created summary sheet '{summary_name}'")

    def save_workbook(self) -> None:
        """
        Save the workbook to the output file.

        Raises:
            PermissionError: If unable to write to the output file
        """
        try:
            # Ensure output directory exists
            self.output_file.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            self.workbook.save(self.output_file)
            logger.info(f"Excel report saved to {self.output_file}")

        except PermissionError:
            logger.error(f"Permission denied: Cannot write to {self.output_file}")
            raise
        except Exception as e:
            logger.error(f"Error saving workbook: {str(e)}")
            raise

    def get_sheet_names(self) -> List[str]:
        """
        Get list of all sheet names in the workbook.

        Returns:
            List of sheet names
        """
        return list(self.sheets.keys())

    def get_sheet_info(self, sheet_name: str) -> Dict[str, Any]:
        """
        Get information about a specific sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            Dictionary with sheet information
        """
        if sheet_name not in self.sheets:
            return {"error": f"Sheet '{sheet_name}' not found"}

        ws = self.sheets[sheet_name]

        return {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "charts": len(ws._charts),
            "conditional_formatting": len(ws.conditional_formatting.cf_rules)
        }
